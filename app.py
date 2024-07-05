import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('Scheduled Hours & Wages.xlsx')
worksheet = workbook.active


import openpyxl
from browser import document, html

def processExcelFile(fileContents):
    # Process the file contents here
    # For example, you can use the openpyxl library to read the Excel file
    workbook = openpyxl.load_workbook(fileContents)
    worksheet = workbook.active

  
employee_names = []
hours_worked = []

# Extract employee names and regular hours from column 5, starting from row 5
for row in worksheet.iter_rows(min_row=5, min_col=5, max_col=5, values_only=True):
    for cell_value in row:
        if cell_value:
            hours_worked.append(float(cell_value))

# Assuming the employee names are in the same order as the hours worked
employee_names = ["Gill, Kuldeep Kaur", "Gujrati, Zeel", "Kaur, Amarjot", "Kaur, Balwinder", "Kaur, Komalpreet", "Kaur, Taranjot", "Maini, Rupam", "Patel, Harisha", "Shah, Priyal", "Shah, Vrushabh", "Sidhu, Jasleen", "Soni, Saloni"]

allowed_hours_on_sin = [0, 40, 60, 60, 60, 20, 0, 40, 20, 20, 60, 50]

pay_period = input("What is the pay period? ")

print(f"\nPay Period: {pay_period}")
print("\n#### Master Chart")
print("{:<20} {:<20} {:<20} {:<20} {:<20}".format("Employee Name", "Hours Worked", "Hours on SIN", "Hours on Cash", "Cash Payout"))

total_cash_payout = 0
for i in range(len(employee_names)):
    hours_on_sin = min(hours_worked[i], allowed_hours_on_sin[i])
    hours_on_cash = max(0, hours_worked[i] - allowed_hours_on_sin[i])
    cash_payout = hours_on_cash * 12
    total_cash_payout += cash_payout
    print("{:<20} {:<20.2f} {:<20.2f} {:<20.2f} ${:<18.2f}".format(employee_names[i], hours_worked[i], hours_on_sin, hours_on_cash, cash_payout))

print(f"\nTotal Cash Payout: ${total_cash_payout:.2f}")

print('/////////////////////////////////////////');
report = "2619959 Ontario Ltd.\nVikas Manocha\n"
for i in range(len(employee_names)):
    if employee_names[i] in ["Kaur, Balwinder", "Patel, Harisha", "Shah, Priyal", "Sidhu, Jasleen", "Soni, Saloni", "Gujrati, Zeel"]:
        hours_to_report = min(hours_worked[i], allowed_hours_on_sin[i])
        report += f"{employee_names[i].split(',')[1].strip()} {hours_to_report} hrs.\n"

report += "\nFor Ganpati International Inc.\nVikas Manocha\n"
for i in range(len(employee_names)):
    if employee_names[i] in ["Kaur, Amarjot", "Kaur, Komalpreet", "Shah, Vrushabh", "Kaur, Taranjot"]:
        hours_to_report = min(hours_worked[i], allowed_hours_on_sin[i])
        report += f"{employee_names[i].split(',')[1].strip()} {hours_to_report} Hrs\n"

print(report)


    # Create HTML elements to display the report
output_div = document["output"]
output_div.innerHTML = ""  # Clear previous content

# Add the pay period
pay_period_heading = html.H4(f"Pay Period: {pay_period}")
output_div <= pay_period_heading

# Add the master chart
master_chart_heading = html.H4("Master Chart")
output_div <= master_chart_heading

table = html.TABLE()
header_row = html.TR()
header_row <= html.TH("Employee Name")
header_row <= html.TH("Hours Worked")
header_row <= html.TH("Hours on SIN")
header_row <= html.TH("Hours on Cash")
header_row <= html.TH("Cash Payout")
table <= header_row

for i in range(len(employee_names)):
    row = html.TR()
    row <= html.TD(employee_names[i])
    row <= html.TD(f"{hours_worked[i]:.2f}")
    row <= html.TD(f"{min(hours_worked[i], allowed_hours_on_sin[i]):.2f}")
    row <= html.TD(f"{max(0, hours_worked[i] - allowed_hours_on_sin[i]):.2f}")
    row <= html.TD(f"${hours_on_cash * 12:.2f}")
    table <= row

output_div <= table

# Add the total cash payout
total_cash_payout_heading = html.H4(f"Total Cash Payout: ${total_cash_payout:.2f}")
output_div <= total_cash_payout_heading

# Add the report
report_heading = html.H4("Report")
output_div <= report_heading
output_div <= html.PRE(report)

# Expose the processExcelFile function to JavaScript
window.pyApp = {
    'processExcelFile': processExcelFile
}